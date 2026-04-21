"""Import members from 'PERSONAL MOSTRADOR.xlsx'.

Parses the 3 sheets (RX, CCEE PLT 0, CCEE PLT 1), detects subsections (Mostrador RX,
Mostrador Laboratorio, CCEE Planta 0, Laboratorio 7h-10h, CCEE Planta 1), and creates:

  - DepartmentMember rows (idempotent by full_name)
  - MemberGenerationPreference rows with shift_rotation / daily_hours / work_days JSONB

The rules column in the Excel (jornada + turnos + días) is translated to a stable
JSON shape the generator can consume. Run from backend/:

    uv run python -m scripts.import_personal_mostrador [path/to/xlsx]
"""

import asyncio
import logging
import platform
import re
import sys
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)-7s | %(message)s")
log = logging.getLogger("horarios.import")

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from src.core.config import settings
from src.models.department_member import DepartmentMember
from src.models.member_generation_preference import MemberGenerationPreference


GROUP_HEADERS: list[tuple[str, str, str]] = [
    # (substring to match in col A, canonical group name, color)
    ("PERSONAL MOSTRADOR RX",         "Mostrador RX",          "#3B82F6"),
    ("PERSONAL   MOSTRADOR RX",       "Mostrador RX",          "#3B82F6"),
    ("MOSTRADOR LABORATORIO",         "Mostrador Laboratorio", "#10B981"),
    ("PERSONAL MOSTRADOR PLANTA 0",   "CCEE Planta 0",         "#F59E0B"),
    ("LABORATORIO 7H A 10H",          "Laboratorio 7h-10h",    "#6366F1"),
    ("PERSONAL PLANTA 1",             "CCEE Planta 1",         "#EC4899"),
]

WEEKDAY_RANGES: dict[str, list[int]] = {
    "lunes a sabado":  [0, 1, 2, 3, 4, 5],
    "lunes a viernes": [0, 1, 2, 3, 4],
    "lunes a domingo": [0, 1, 2, 3, 4, 5, 6],
}

DEFAULT_WORK_DAYS = [0, 1, 2, 3, 4]  # Mon-Fri fallback


def _norm(v) -> str:
    return (str(v) if v is not None else "").strip()


def _match_group(cell_a: str) -> tuple[str, str] | None:
    up = cell_a.upper()
    for needle, name, color in GROUP_HEADERS:
        if needle in up:
            return name, color
    return None


def _parse_rule(jornada: str, turnos: str, dias: str) -> dict:
    """Translate (jornada, turnos, dias) → generator-ready JSONB.

    Examples:
        ("rotatorio 7h", "M/T", "lunes a sabado")
          -> {"shift_rotation": ["M","T"], "daily_hours": 7, "work_days": [0..5]}
        ("M 8 horas", "M", "lunes a viernes")
          -> {"shift_rotation": ["M"],     "daily_hours": 8, "work_days": [0..4]}
        ("reduccion rotatorio 5 H", "R/ M-T", "lunes a viernes")
          -> {"shift_rotation": ["M","T"], "daily_hours": 5, "work_days": [0..4]}
    """
    j = jornada.lower()
    t = turnos.upper().replace(" ", "").replace("R/", "")  # strip reduccion prefix
    d = dias.lower().strip()

    # daily_hours: first integer in jornada string (supports "5H", "7h", "8 horas")
    m = re.search(r"(\d+)", j)
    daily_hours = int(m.group(1)) if m else 8

    # shift_rotation: prefer "turnos" column; fall back to jornada hints
    if t:
        parts = re.split(r"[/\-]", t)
        rotation = [p for p in parts if p in {"M", "T", "N"}]
    else:
        rotation = []
    if not rotation:
        # Last-resort inference from the jornada label
        if "m " in f" {j} " or j.startswith("m"):
            rotation = ["M"]
        else:
            rotation = ["M", "T"]

    # work_days: match any known range substring in "dias"
    work_days = None
    for needle, days in WEEKDAY_RANGES.items():
        if needle in d:
            work_days = days
            break
    if work_days is None:
        work_days = DEFAULT_WORK_DAYS

    return {
        "shift_rotation": rotation,
        "daily_hours": daily_hours,
        "work_days": work_days,
    }


def parse_workbook(xlsx_path: Path) -> list[dict]:
    """Return a flat list of {full_name, role_name, group_name, color, rule_jsonb}."""
    wb = load_workbook(str(xlsx_path), data_only=True)
    out: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_group: str = sheet_name
        current_color: str = "#6B7280"

        for row in ws.iter_rows(values_only=True):
            cell_a = _norm(row[0])
            if not cell_a:
                continue

            group = _match_group(cell_a)
            if group:
                current_group, current_color = group
                continue

            # Heuristic: a member row has jornada (col B) filled
            jornada = _norm(row[1]) if len(row) > 1 else ""
            if not jornada:
                continue

            turnos = _norm(row[2]) if len(row) > 2 else ""
            dias = _norm(row[3]) if len(row) > 3 else ""

            rule = _parse_rule(jornada, turnos, dias)
            weekly_limit = rule["daily_hours"] * min(5, len(rule["work_days"]))

            out.append({
                "full_name": cell_a,
                "role_name": jornada or "Personal",
                "group_name": current_group,
                "color_tag": current_color,
                "weekly_hour_limit": float(weekly_limit),
                "rule_jsonb": rule,
            })

    return out


async def upsert(session: AsyncSession, rows: list[dict]) -> tuple[int, int]:
    created = updated = 0

    for r in rows:
        existing = await session.execute(
            select(DepartmentMember).where(DepartmentMember.full_name == r["full_name"])
        )
        member = existing.scalar_one_or_none()

        if member is None:
            member = DepartmentMember(
                full_name=r["full_name"],
                role_name=r["role_name"],
                weekly_hour_limit=r["weekly_hour_limit"],
                color_tag=r["color_tag"],
                group_name=r["group_name"],
            )
            session.add(member)
            await session.flush()
            created += 1
            log.info("created  %-40s  [%s]", r["full_name"], r["group_name"])
        else:
            member.role_name = r["role_name"]
            member.weekly_hour_limit = r["weekly_hour_limit"]
            member.color_tag = r["color_tag"]
            member.group_name = r["group_name"]
            updated += 1
            log.info("updated  %-40s  [%s]", r["full_name"], r["group_name"])

        pref_row = await session.execute(
            select(MemberGenerationPreference).where(MemberGenerationPreference.member_id == member.id)
        )
        pref = pref_row.scalar_one_or_none()
        if pref is None:
            session.add(MemberGenerationPreference(member_id=member.id, preferences_jsonb=r["rule_jsonb"]))
        else:
            pref.preferences_jsonb = r["rule_jsonb"]

    await session.commit()
    return created, updated


async def main(xlsx_path: Path) -> None:
    rows = parse_workbook(xlsx_path)
    log.info("Parsed %d members from %s", len(rows), xlsx_path.name)

    engine = create_async_engine(settings.get_database_url())
    async_session = async_sessionmaker(engine, expire_on_commit=False)
    async with async_session() as session:
        created, updated = await upsert(session, rows)
    await engine.dispose()

    log.info("Done: %d created, %d updated", created, updated)


if __name__ == "__main__":
    if platform.system() == "Windows":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    default = Path(__file__).resolve().parent.parent / "PERSONAL MOSTRADOR.xlsx"
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else default
    if not xlsx.exists():
        log.error("File not found: %s", xlsx)
        sys.exit(1)

    asyncio.run(main(xlsx))
