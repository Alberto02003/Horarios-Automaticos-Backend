"""Learn per-member weekly patterns from a historical schedule xlsx.

Input: a monthly schedule export like `horarios_Abr_2026.xlsx` — first column is
member name, header row uses labels like "Mie\\n1" / "Jue\\n2" / ... so each
column's weekday is read from the Spanish prefix.

For each member, we compute the most common shift code per weekday across all
weeks of the sheet and write it to `member_generation_preferences.preferences_jsonb`
as `weekly_pattern` (7-element list, Mon..Sun).

Run from backend/:

    uv run python -m scripts.learn_patterns_from_xlsx ../horarios_Abr_2026.xlsx

The script is idempotent — overwrites the weekly_pattern for existing members,
leaves other JSONB keys untouched, skips members not found in the DB.
"""

import asyncio
import logging
import platform
import sys
from collections import Counter
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)-7s | %(message)s")
log = logging.getLogger("horarios.learn_patterns")

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from src.core.config import settings
from src.models.department_member import DepartmentMember
from src.models.member_generation_preference import MemberGenerationPreference


WEEKDAY_ES = {
    "LUN": 0, "MAR": 1, "MIE": 2, "MIÉ": 2,
    "JUE": 3, "VIE": 4, "SAB": 5, "SÁB": 5, "DOM": 6,
}


def parse_header(header_row: tuple) -> list[int]:
    """Return weekday index (0=Mon..6=Sun) for each data column after the name."""
    weekdays = []
    for cell in header_row[1:]:
        if cell is None:
            weekdays.append(None)
            continue
        prefix = str(cell).split("\n", 1)[0].strip().upper()
        weekdays.append(WEEKDAY_ES.get(prefix))
    return weekdays


def learn_from_rows(rows: list[tuple]) -> dict[str, list[str | None]]:
    """Return {member_name -> [shift_code or None] * 7} inferring the mode per weekday."""
    weekdays = parse_header(rows[0])
    result: dict[str, list[str | None]] = {}

    for row in rows[1:]:
        name = row[0]
        if not name:
            continue
        name = str(name).strip()

        counters: list[Counter] = [Counter() for _ in range(7)]
        for wd, cell in zip(weekdays, row[1:1 + len(weekdays)]):
            if wd is None or cell is None:
                continue
            code = str(cell).strip()
            if code:
                counters[wd][code] += 1

        pattern: list[str | None] = []
        for wd, c in enumerate(counters):
            if not c:
                pattern.append(None)
            else:
                pattern.append(c.most_common(1)[0][0])
        result[name] = pattern

    return result


def learn_from_workbook(xlsx_path: Path) -> dict[str, list[str | None]]:
    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    return learn_from_rows(rows)


async def save_patterns(patterns: dict[str, list[str | None]]) -> tuple[int, int, int]:
    engine = create_async_engine(settings.get_database_url())
    async_session = async_sessionmaker(engine, expire_on_commit=False)
    updated = created = skipped = 0

    async with async_session() as session:
        for name, pattern in patterns.items():
            result = await session.execute(
                select(DepartmentMember).where(DepartmentMember.full_name == name)
            )
            member = result.scalar_one_or_none()
            if member is None:
                log.warning("miembro no encontrado: %s", name)
                skipped += 1
                continue

            pref_result = await session.execute(
                select(MemberGenerationPreference).where(
                    MemberGenerationPreference.member_id == member.id
                )
            )
            pref = pref_result.scalar_one_or_none()
            if pref is None:
                session.add(MemberGenerationPreference(
                    member_id=member.id,
                    preferences_jsonb={"weekly_pattern": pattern},
                ))
                created += 1
            else:
                data = dict(pref.preferences_jsonb or {})
                data["weekly_pattern"] = pattern
                pref.preferences_jsonb = data
                updated += 1
            log.info("%-40s  pattern=%s", name[:40], pattern)

        await session.commit()

    await engine.dispose()
    return created, updated, skipped


async def main(xlsx_path: Path) -> None:
    patterns = learn_from_workbook(xlsx_path)
    log.info("Aprendidos %d patrones desde %s", len(patterns), xlsx_path.name)
    created, updated, skipped = await save_patterns(patterns)
    log.info("Hecho: %d creados, %d actualizados, %d saltados", created, updated, skipped)


if __name__ == "__main__":
    if platform.system() == "Windows":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    if len(sys.argv) < 2:
        default = Path(__file__).resolve().parent.parent.parent / "horarios_Abr_2026.xlsx"
        if not default.exists():
            log.error("Pasa la ruta al xlsx como primer argumento")
            sys.exit(1)
        xlsx = default
    else:
        xlsx = Path(sys.argv[1])

    if not xlsx.exists():
        log.error("File not found: %s", xlsx)
        sys.exit(1)

    asyncio.run(main(xlsx))
