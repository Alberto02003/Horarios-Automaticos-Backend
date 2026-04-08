"""Export schedules to Excel."""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.schedule_assignment import ScheduleAssignment
from src.models.department_member import DepartmentMember
from src.models.shift_type import ShiftType
from src.models.schedule_period import SchedulePeriod


DAY_NAMES = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]


def _hex_to_argb(hex_color: str) -> str:
    return "FF" + hex_color.lstrip("#")


async def export_excel(db: AsyncSession, period_id: int) -> bytes:
    # Load data
    period_result = await db.execute(select(SchedulePeriod).where(SchedulePeriod.id == period_id))
    period = period_result.scalar_one()

    members_result = await db.execute(select(DepartmentMember).where(DepartmentMember.is_active.is_(True)).order_by(DepartmentMember.full_name))
    members = list(members_result.scalars().all())

    shifts_result = await db.execute(select(ShiftType))
    shifts = {s.id: s for s in shifts_result.scalars().all()}

    assignments_result = await db.execute(
        select(ScheduleAssignment).where(ScheduleAssignment.schedule_period_id == period_id)
    )
    assignment_map: dict[tuple[int, date], ScheduleAssignment] = {}
    for a in assignments_result.scalars().all():
        assignment_map[(a.member_id, a.date)] = a

    # Build dates
    dates: list[date] = []
    current = period.start_date
    from datetime import timedelta
    while current <= period.end_date:
        dates.append(current)
        current += timedelta(days=1)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = period.name

    thin_border = Border(
        left=Side(style="thin", color="FFCCCCCC"),
        right=Side(style="thin", color="FFCCCCCC"),
        top=Side(style="thin", color="FFCCCCCC"),
        bottom=Side(style="thin", color="FFCCCCCC"),
    )

    # Header row: "Miembro" + day numbers
    header_fill = PatternFill(start_color="FFF9A8D4", end_color="FFF9A8D4", fill_type="solid")
    ws.cell(row=1, column=1, value="Miembro").font = Font(bold=True, size=10)
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = thin_border
    ws.column_dimensions["A"].width = 22

    for i, d in enumerate(dates):
        col = i + 2
        day_name = DAY_NAMES[d.weekday()]
        cell = ws.cell(row=1, column=col, value=f"{day_name}\n{d.day}")
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

        is_weekend = d.weekday() >= 5
        cell.fill = PatternFill(start_color="FFFCE7F3", end_color="FFFCE7F3", fill_type="solid") if is_weekend else header_fill
        ws.column_dimensions[get_column_letter(col)].width = 5.5

    # Data rows
    for row_idx, member in enumerate(members, start=2):
        name_cell = ws.cell(row=row_idx, column=1, value=member.full_name)
        name_cell.font = Font(size=9)
        name_cell.border = thin_border

        for i, d in enumerate(dates):
            col = i + 2
            assignment = assignment_map.get((member.id, d))
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            if assignment:
                shift = shifts.get(assignment.shift_type_id)
                if shift:
                    cell.value = shift.code
                    cell.font = Font(bold=True, size=9, color="FFFFFFFF")
                    cell.fill = PatternFill(start_color=_hex_to_argb(shift.color), end_color=_hex_to_argb(shift.color), fill_type="solid")

    # Freeze first column and header
    ws.freeze_panes = "B2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
